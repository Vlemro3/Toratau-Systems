"""
Чтение файлов смет (.xls, .xlsx, .xml) и извлечение позиций в табличную часть.
Без нейросети — прямое чтение структуры файла.
"""
from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

ALLOWED_EXTENSIONS = ("xls", "xlsx", "xml")


def _safe_float(val: Any, default: float = 0) -> float:
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip().replace(",", ".").replace("\xa0", "").replace(" ", "")
        return float(s) if s else default
    except (TypeError, ValueError):
        return default


def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _normalize_position(p: dict[str, Any], i: int) -> dict[str, Any]:
    return {
        "id": i + 1,
        "num": _safe_str(p.get("num", p.get("number", i + 1))),
        "normCode": _safe_str(p.get("normCode", p.get("code", ""))),
        "name": _safe_str(p.get("name", p.get("description", ""))),
        "unit": _safe_str(p.get("unit", p.get("measure", ""))),
        "volume": _safe_float(p.get("volume", p.get("quantity", p.get("qty")))),
        "price": _safe_float(p.get("price", p.get("unitPrice"))),
        "total": _safe_float(p.get("total", p.get("sum", p.get("amount")))),
        "overhead": _safe_float(p.get("overhead", p.get("nr"))),
        "profit": _safe_float(p.get("profit", p.get("sp"))),
        "adjustmentCoeff": _safe_str(p.get("adjustmentCoeff", "")),
        "recalcCoeffNumber": _safe_str(p.get("recalcCoeffNumber", "")),
        "laborPersonHours": _safe_float(p.get("laborPersonHours", p.get("labor"))),
        "costPerUnitFromStart": _safe_float(p.get("costPerUnitFromStart", "")),
    }


def parse_estimate_file(content: bytes, filename: str, base_type: str = "FER") -> list[dict[str, Any]]:
    """
    Читает файл сметы (.xls, .xlsx, .xml) и возвращает список позиций.
    PDF не поддерживается.
    """
    ext = (filename or "").lower().split(".")[-1].strip()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(
            f"Неподдерживаемый формат файла. Разрешены только: .xls, .xlsx, .xml (получен .{ext or 'unknown'})"
        )

    if ext == "xml":
        positions = _parse_xml(content)
    elif ext == "xlsx":
        positions = _parse_xlsx(content)
    else:
        positions = _parse_xls(content)

    if not positions:
        raise ValueError("В файле не найдено ни одной позиции сметы. Проверьте структуру файла.")

    return [_normalize_position(p, i) for i, p in enumerate(positions)]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        return []
    rows = list(ws.iter_rows(values_only=True))
    return _excel_rows_to_positions(rows)


def _parse_xls(content: bytes) -> list[dict[str, Any]]:
    import xlrd
    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    return _excel_rows_to_positions(rows)


# Варианты заголовков для определения колонок (нижний регистр)
HEADER_ALIASES = {
    "num": ("№ п/п", "№п/п", "№", "номер", "num", "number", "n п/п", "поз", "позиция"),
    "normCode": ("обоснование", "шифр", "код", "норма", "норм", "code", "norm", "фер", "тер", "гесн"),
    "name": ("наименование работ", "наименование", "название", "описание", "name", "description", "работы и затрат"),
    "unit": ("единица измерения", "ед.изм", "ед. изм", "ед", "единица", "unit", "measure", "измерен"),
    "volume": ("на единицу измерения", "кол-во", "кол", "количество", "объем", "volume", "quantity", "qty", "объём"),
    "price": ("цена", "цена ед", "price", "unitprice", "руб", "на единицу измерения в базовом уровне цен"),
    "total": ("всего затрат", "всего в текущих", "всего", "сумма", "total", "sum", "amount", "итого", "сметная стоимость"),
    "overhead": ("нр", "накладные", "overhead"),
    "profit": ("сп", "прибыль", "profit"),
}

# Ключевые слова, которые однозначно определяют строку как строку заголовков таблицы
_HEADER_KEYWORDS = ("наименование", "название", "обоснование", "ед.изм", "единица", "name", "code", "№ п/п", "№п/п")
# Строки-разделители, которые не являются позициями (секции, разделы и т.п.)
_SKIP_PREFIXES = ("итого", "всего", "total", "сумма", "раздел", "глава", "в том числе", "накладные расходы", "сметная прибыль")


def _find_header_row(rows: list[tuple[Any, ...] | list[Any]]) -> tuple[int, dict[str, int]]:
    """Ищет строку заголовков по всему файлу, возвращает (индекс_строки, col_map)."""
    best_row = -1
    best_map: dict[str, int] = {}

    for row_idx, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(cells)

        # Проверяем, содержит ли строка ключевые слова заголовков
        if not any(kw in joined for kw in _HEADER_KEYWORDS):
            continue

        col_map: dict[str, int] = {}
        for key, aliases in HEADER_ALIASES.items():
            for idx, cell in enumerate(cells):
                if not cell:
                    continue
                cell_clean = re.sub(r"[^\wа-яё/]", "", cell, flags=re.I)
                for al in aliases:
                    al_lower = al.lower()
                    al_clean = re.sub(r"[^\wа-яё/]", "", al_lower, flags=re.I)
                    if al_lower in cell or (len(cell_clean) >= 2 and al_clean and al_clean in cell_clean):
                        col_map[key] = idx
                        break
                if key in col_map:
                    break

        # Выбираем строку с максимальным числом совпавших колонок
        if len(col_map) > len(best_map):
            best_map = col_map
            best_row = row_idx

    return best_row, best_map


def _excel_rows_to_positions(rows: list[tuple[Any, ...] | list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    # Ищем строку заголовков по всему файлу
    header_idx, col_map = _find_header_row(rows)

    # Если заголовки не найдены — пробуем fallback по умолчанию
    if not col_map:
        col_map = {
            "num": 0, "normCode": 1, "name": 2, "unit": 3,
            "volume": 4, "price": 5, "total": 6, "overhead": 7, "profit": 8,
        }
        header_idx = -1

    start = header_idx + 1 if header_idx >= 0 else 0

    # Пропускаем строки-подзаголовки сразу после заголовка (например, строку с нумерацией "1 2 3 4 ...")
    if start < len(rows):
        check_row = [str(c).strip() if c is not None else "" for c in rows[start]]
        if all(re.match(r"^\d{1,2}$", v) for v in check_row if v):
            start += 1

    result = []
    for row in rows[start:]:
        if not isinstance(row, (list, tuple)):
            continue
        arr = list(row) if isinstance(row, tuple) else row
        if not arr or all(v is None or (isinstance(v, str) and not str(v).strip()) for v in arr):
            continue

        # Пропускаем строки-разделители
        first_cell = str(arr[0]).strip().lower() if arr[0] is not None else ""
        # Проверяем первые непустые ячейки на наличие «раздел», «итого» и т.д.
        row_text = " ".join(str(c).strip().lower() for c in arr[:4] if c is not None)
        if any(row_text.startswith(sp) for sp in _SKIP_PREFIXES):
            continue
        if first_cell in ("итого", "всего", "total", "сумма") or first_cell.startswith("итого"):
            continue
        # Пропускаем строки «Раздел N» или текстовые заголовки секций
        if re.match(r"^раздел\b", first_cell):
            continue

        p: dict[str, Any] = {}
        for key, idx in col_map.items():
            if idx < len(arr) and arr[idx] is not None:
                p[key] = arr[idx]

        # Позиция валидна, если есть хотя бы имя/код и числовые значения
        has_name = bool(p.get("name") or p.get("normCode"))
        has_numbers = _safe_float(p.get("volume")) > 0 or _safe_float(p.get("price")) > 0 or _safe_float(p.get("total")) > 0
        if has_name and has_numbers:
            if "total" not in p and ("volume" in p or "price" in p):
                p["total"] = _safe_float(p.get("volume")) * _safe_float(p.get("price"))
            result.append(p)
    return result


def _parse_xml(content: bytes) -> list[dict[str, Any]]:
    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        raise ValueError("Некорректный XML в файле")
    positions = []
    # Типичные корневые элементы с позициями в сметах (ГЭСН/ФЕР/ТЕР экспорт)
    for tag in ("Positions", "positions", "Items", "items", "Rows", "rows", "Data", "Estimate", "Смета"):
        items = root.findall(f".//{tag}") or root.findall(f".//{tag.lower()}")
        for parent in items:
            for row in parent:
                p = _xml_row_to_position(row)
                if p:
                    positions.append(p)
    if not positions:
        # Плоский список: все элементы с дочерними полями
        for row in root.iter():
            if len(row) >= 3 and row.tag not in ("Positions", "positions", "Items", "items"):
                p = _xml_row_to_position(row)
                if p:
                    positions.append(p)
    return positions


def _xml_row_to_position(elem: ET.Element) -> dict[str, Any] | None:
    by_name: dict[str, str] = {}
    for child in elem:
        local = child.tag.split("}")[-1] if "}" in child.tag else child.tag
        text = (child.text or "").strip()
        if len(child):
            text = " ".join((t or "").strip() for t in child.itertext())
        by_name[local.lower()] = text
    num = by_name.get("num") or by_name.get("number") or by_name.get("n") or by_name.get("pos") or ""
    name = by_name.get("name") or by_name.get("description") or by_name.get("naimenovanie") or by_name.get("title") or ""
    code = by_name.get("code") or by_name.get("normcode") or by_name.get("norm") or ""
    unit = by_name.get("unit") or by_name.get("measure") or by_name.get("edizm") or ""
    if not name and not code and not num:
        return None
    volume = by_name.get("volume") or by_name.get("quantity") or by_name.get("qty") or by_name.get("kol") or ""
    price = by_name.get("price") or by_name.get("unitprice") or by_name.get("cenа") or ""
    total = by_name.get("total") or by_name.get("sum") or by_name.get("amount") or by_name.get("summa") or ""
    overhead = by_name.get("overhead") or by_name.get("nr") or ""
    profit = by_name.get("profit") or by_name.get("sp") or ""
    return {
        "num": num,
        "normCode": code,
        "name": name,
        "unit": unit,
        "volume": volume,
        "price": price,
        "total": total,
        "overhead": overhead,
        "profit": profit,
    }
